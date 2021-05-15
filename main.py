import requests
from bs4 import BeautifulSoup
import os
import openpyxl
import re
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment

path='./豆瓣'
rank=0
#初始化一张工作表
def start_excel(book,sh):
    # 修改当前sheet标题
    sh.title = '豆瓣TOP250'
    row = ['排名','电影','电影别称','电影类型','电影评分']
    rows = ['排名', '电影', '电影别称', '电影类型', '电影评分','sssssssssssssss']
    sh.append(row)
    book.save('豆瓣TOP250.xlsx')
    print('创建工作表完成！')

#获取电影封面
def get_img_movie(url,soup,len_movies):
    for i in range(0, len_movies):
        # 每个电影名称
        img_name = soup.select('#wrapper .grid_view > li img')[i]['alt']
        path_img = os.getcwd() + '/' + str(i + 1) + img_name + '.jpg'
        # 电影图片
        url_img_movie = soup.select('#wrapper .grid_view > li img')[i]['src']
        # 获取图片数据并以二进制保存在img_movie里
        img_movie = requests.get(url=url_img_movie, headers=header).content
        # 存储
        with open(path_img, 'wb') as fp:
            fp.write(img_movie)
        print(img_name + '已下载完成!')


#电影信息
def get_movie_message(soup,len_movies,book,sh):
    #影名匹配模式
    re_mov_name = '<span class="title">(.*?)</span>'
    #信息匹配模式
    re_messa_movie = '<br>(.*?[\n]*?.*?[\n]*?.*?)</p>'
    #其他影名匹配模式
    re_oth_name = '<span class="other">(.*?)</span>'
    # 电影评分匹配模式
    re_score = '<span class="rating_num" property="v:average">([0-9]\.[0-9])</span>'

    for i in range(0, len_movies):
        #添加的内容
        list_movie = []
        global rank
        rank += 1
        list_movie.append(rank)
        mov_name1 = soup.select('#wrapper .grid_view > li')[i]
        temp_mov_name = re.findall(pattern=re_mov_name, string=str(mov_name1))
        # 拼接有多个名字的电影
        if len(temp_mov_name) > 1:
            mov_name = ''
            for a in range(0, len(temp_mov_name)):
                mov_name += temp_mov_name[a]
        else:
            mov_name = temp_mov_name[0]
        list_movie.append(mov_name)
        #print(mov_name)
        # 电影的其他名字
        mov_oth_name = soup.select('#wrapper .grid_view > li .other')[i]
        mov_oth_name = re.findall(re_oth_name,str(mov_oth_name))
        list_movie.append(mov_oth_name[0])
        #print(mov_oth_name)
        #电影类型(待用正则优化)
        type_movie = soup.select('#wrapper .grid_view > li')[i]
        num_type_movie = str(type_movie).find('<br/>')
        num_type_movie1 = str(type_movie).find('</p>')
        temp_num_type_movie = str(type_movie)[num_type_movie:num_type_movie1]
        list_movie.append(temp_num_type_movie[5:-6].strip())
        #print(temp_num_type_movie[5:-6].strip())
        # 电影评分
        score_movie = soup.select('#wrapper .grid_view > li .star')[i]
        score_movie = re.findall(re_score,str(score_movie))
        list_movie.append(score_movie[0])
        print(list_movie)
        #将数据导入excel
        sh.append(list_movie)
        #print(score_movie)
        book.save('豆瓣TOP250.xlsx')




if __name__ == '__main__':
    #网页头信息
    headers={
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.128 Safari/537.36',
        'referer': 'https: // pagead2.googlesyndication.com /'
    }
    #图片头信息
    header = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.128 Safari/537.36',
        'Upgrade-Insecure-Requests': '1',
        'sec - ch - ua - mobile': '?0',
        'sec - ch - ua': '"Google Chrome";v = "89", "Chromium";v = "89", ";Not A Brand";v = "99"'
    }
    url = 'https://movie.douban.com/top250?start={}&filter='
    # 创建程序所在文件夹
    if not os.path.isdir(path):
        os.mkdir(path)
    # 将工作目录改为刚创建的文件夹里
    os.chdir(path)

    # 创建一张excel对象
    book = openpyxl.Workbook()
    # 创建新的工作簿
    sh = book.active
    # 初始化一张工作表
    start_excel(book=book, sh=sh)
    #页数
    pagenum=-25
    for a in range(0,10):
        pagenum+=25
        #第n页的新网址
        new_url=url.format(pagenum)
        #获取整个网页源码
        page_data = requests.get(url=new_url,headers=headers).text
        #煲汤实例化对象
        soup = BeautifulSoup(page_data,'lxml')
        #寻找各个电影封面
        len_movies = len(soup.select('#wrapper .grid_view > li img'))
        #收集每部电影的封面
        get_img_movie(url=new_url,soup=soup,len_movies=len_movies)
        # 电影信息
        get_movie_message(soup=soup,len_movies=len_movies,book=book,sh=sh)
    for i in range(1, 6):
        sh.cell(1, i).font = Font(size=16)
    # 设置行高
    for i in range(1,251):
        sh.row_dimensions[i].height = 25
    # 设置列宽
    sh.row_dimensions[1].width = 5
    for i in range(66,70):
        sh.column_dimensions[chr(i)].width = 30
    # 通过参数horizontal和vertical来设置文字在单元格里的对齐方式，此外设置值还可为left和right
    for i in range(1,6):
        for j in range(1,252):
            sh.cell(j,i).alignment = Alignment(horizontal='center', vertical='center')
    book.save('豆瓣TOP250.xlsx')

    print('爬取完成！')


